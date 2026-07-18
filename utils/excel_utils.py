# NEW ARCHITECTURE (Use this today)
import openpyxl
import os


def get_data_from_excel(file_name, sheet_name):
    """
    Reads an Excel sheet and returns the data as a list of tuples.
    Perfect for @pytest.mark.parametrize.
    """
    # 1. Dynamically build the path to the testdata folder
    base_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    excel_path = os.path.join(base_path, "test_data", file_name)

    # 2. Open the workbook EXACTLY ONCE
    workbook = openpyxl.load_workbook(excel_path)
    print(workbook.sheetnames)
    sheet = workbook[sheet_name]

    data_list = []

    # 3. Read the data (skipping the header row)
    for row in range(2, sheet.max_row + 1):
        row_data = []
        for col in range(1, sheet.max_column + 1):
            row_data.append(sheet.cell(row=row, column=col).value)

        data_list.append(tuple(row_data))

    return data_list